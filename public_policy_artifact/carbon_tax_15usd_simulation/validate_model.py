from __future__ import annotations

import re
import sys
from pathlib import Path

from openpyxl import load_workbook


EXPECTED_SHEETS = [
    "Cover", "Executive Summary", "Scenario Comparison", "Decision Matrix",
    "Assumptions Log", "Revenue Build-Up", "Cost & Investment", "P&L Waterfall",
    "ROI & Payback", "Sensitivity Analysis", "Data Sources", "Methodology Notes",
    "Sector Incidence",
]
ERROR_TOKENS = {"#REF!", "#DIV/0!", "#VALUE!", "#N/A", "#NAME?", "#NUM!", "#NULL!"}


def named_value(wb_formula, wb_values, name: str):
    dn = wb_formula.defined_names[name]
    sheet, coord = next(dn.destinations)
    return wb_values[sheet][coord].value


def main(path: str) -> int:
    p = Path(path)
    wf = load_workbook(p, data_only=False)
    wv = load_workbook(p, data_only=True)
    errors: list[str] = []
    warnings: list[str] = []

    if wf.sheetnames != EXPECTED_SHEETS:
        errors.append(f"Sheet order mismatch: {wf.sheetnames}")

    formula_count = 0
    for ws in wf.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formula_count += 1
                    if any(token in cell.value for token in ERROR_TOKENS):
                        errors.append(f"Formula contains error token: {ws.title}!{cell.coordinate}")
                cached = wv[ws.title][cell.coordinate].value
                if isinstance(cached, str) and cached in ERROR_TOKENS:
                    errors.append(f"Cached Excel error: {ws.title}!{cell.coordinate} = {cached}")

    for name, dn in wf.defined_names.items():
        try:
            destinations = list(dn.destinations)
        except Exception as exc:
            errors.append(f"Named range {name} does not resolve: {exc}")
            continue
        if len(destinations) != 1:
            errors.append(f"Named range {name} has {len(destinations)} destinations")
            continue
        sheet, coord = destinations[0]
        if sheet not in wf.sheetnames or not re.fullmatch(r"\$?[A-Z]{1,3}\$?\d+", coord):
            errors.append(f"Named range {name} has invalid target {sheet}!{coord}")

    prob_sum = sum(named_value(wf, wv, f"scenario_{s}_probability") for s in ("base", "bull", "bear"))
    if abs(prob_sum - 1.0) > 1e-9:
        errors.append(f"Scenario probabilities sum to {prob_sum:.6f}, not 1.0")

    thesis = named_value(wf, wv, "tax_only_base_emissions_reduction_pct")
    if thesis is None or abs(thesis - 0.029) > 1e-9:
        errors.append(f"Thesis replication failed: {thesis}")

    critical = [
        "tax_only_base_emissions_reduction_pct", "tax_only_base_net_gdp_impact_pct",
        "just_transition_base_emissions_reduction_pct", "just_transition_base_net_gdp_impact_pct",
        "accelerated_base_emissions_reduction_pct", "accelerated_base_ndc_progress_pct",
    ]
    for name in critical:
        if named_value(wf, wv, name) is None:
            errors.append(f"Missing cached value after recalculation: {name}")

    if formula_count < 500:
        errors.append(f"Unexpectedly low formula count: {formula_count}")
    if p.stat().st_size > 5_000_000:
        warnings.append(f"Workbook size exceeds 5MB: {p.stat().st_size}")

    print(f"Workbook: {p}")
    print(f"Sheets: {len(wf.sheetnames)}")
    print(f"Named ranges: {len(wf.defined_names)}")
    print(f"Formula cells: {formula_count}")
    print(f"Scenario probability sum: {prob_sum:.0%}")
    print(f"$15 tax-only emissions reduction: {thesis:.1%}")
    print(f"Errors: {len(errors)}")
    for item in errors:
        print(f"ERROR: {item}")
    print(f"Warnings: {len(warnings)}")
    for item in warnings:
        print(f"WARNING: {item}")
    print("RESULT: PASS" if not errors else "RESULT: FAIL")
    return 0 if not errors else 1


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1]))

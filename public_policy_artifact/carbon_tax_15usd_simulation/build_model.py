from __future__ import annotations

import csv
from datetime import date
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.comments import Comment
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import absolute_coordinate, get_column_letter, quote_sheetname
from openpyxl.workbook.defined_name import DefinedName


ROOT = Path(__file__).resolve().parent
OUTPUT = ROOT / "vietnam_carbon_tax_15usd_impact_model_v1_20260714.xlsx"
DATA_INVENTORY = ROOT / "data_inventory.csv"

GREEN = "00B14F"
DARK_GREEN = "006850"
DARK = "1F2227"
GRAY = "404040"
MID_GRAY = "7F7F7F"
LIGHT_GRAY = "E7E6E6"
BLUE = "2C6FA6"
INPUT_BLUE = "DCE6F1"
INPUT_FONT = "1F5C9E"
RESEARCH_GREEN = "D9EAD3"
RESEARCH_FONT = "375623"
AMBER = "D4891A"
YELLOW = "FFF2CC"
RED = "C00000"
WHITE = "FFFFFF"
BLACK = "000000"

thin = Side(style="thin", color="BFBFBF")
medium = Side(style="medium", color="000000")
total_border = Border(top=medium, bottom=thin)
section_border = Border(bottom=thin)

URL_ARTICLE = "https://vnexpress.net/ap-thue-15-usd-moi-tan-co2-co-the-giup-viet-nam-giam-3-phat-thai-5096991.html"
URL_NDC = "https://unfccc.int/sites/default/files/NDC/2022-11/Viet%20Nam_NDC_2022_Eng.pdf"
URL_WB_GDP = "https://data.worldbank.org/country/viet-nam?locations=VN&name_desc=false"
URL_WB_2045 = "https://www.worldbank.org/en/country/vietnam/publication/viet-nam-2045-growing-greener-pathways-to-a-resilient-and-sustainable-future"
URL_OECD_EFFECT = "https://www.oecd.org/en/publications/estimating-the-co2-emission-and-revenue-effects-of-carbon-pricing_39aa16d4-en.html"


ASSUMPTIONS = [
    # name, category, value, unit, source_type, source_title, url, confidence, notes, key_swing
    ("Baseline GHG emissions", "Market", 927.9, "MtCO2e", "Research", "Viet Nam NDC 2022: 2030 BAU", URL_NDC, "High", "2030 BAU projection; used to scale absolute emissions outputs.", False),
    ("Reference nominal GDP", "Macro", 476.39, "USD bn", "Research", "World Bank: Viet Nam GDP 2024", URL_WB_GDP, "High", "Used only to translate percentage GDP effects into an indicative USD scale.", False),
    ("Anchor carbon price", "Policy", 15.0, "USD/tCO2e", "Research", "VnExpress summary of IPSS-UNDP study", URL_ARTICLE, "Medium", "Calibration point reported by article.", True),
    ("Anchor emissions reduction", "Emissions", 0.029, "% of BAU", "Research", "VnExpress summary of IPSS-UNDP study", URL_ARTICLE, "Medium", "2.9% reduction at USD15/tCO2e.", True),
    ("Anchor GDP drag", "Macro", 0.008, "% of GDP", "Research", "VnExpress summary of IPSS-UNDP study", URL_ARTICLE, "Medium", "GDP could decline 0.8% without suitable offsets.", True),
    ("Anchor trade drag", "Macro", 0.009, "% vs baseline", "Research", "VnExpress summary of IPSS-UNDP study", URL_ARTICLE, "Medium", "Exports-imports reported down around 0.9%.", False),
    ("Anchor investment drag", "Macro", 0.013, "% vs baseline", "Research", "VnExpress summary of IPSS-UNDP study", URL_ARTICLE, "Medium", "Investment reported down around 1.3%.", False),
    ("Anchor government revenue index uplift", "Fiscal", 0.0041, "% vs baseline", "Research", "VnExpress summary of IPSS-UNDP study", URL_ARTICLE, "Low", "Article says government revenue rises 0.41%; interpreted as a relative index, not percentage points of GDP.", False),
    ("Revenue recycling threshold", "Equity", 0.30, "% of collected revenue", "Research", "VnExpress summary of IPSS-UNDP study", URL_ARTICLE, "Medium", "30% of proceeds can compensate the lowest-income 50% of households.", True),
    ("Bottom households potentially compensated", "Equity", 0.50, "% of households", "Research", "VnExpress summary of IPSS-UNDP study", URL_ARTICLE, "Medium", "Coverage anchor at the 30% recycling threshold.", False),
    ("Efficiency benchmark", "Policy", 0.10, "% energy efficiency gain", "Research", "VnExpress summary of IPSS-UNDP study", URL_ARTICLE, "Medium", "10% energy saving/efficiency package benchmark.", True),
    ("Efficiency GDP benefit at benchmark", "Macro", 0.01, "% of GDP", "Research", "VnExpress summary of IPSS-UNDP study", URL_ARTICLE, "Medium", "Study reports 10% efficiency could raise GDP by 1%.", True),
    ("NDC unconditional reduction", "Target", 146.3, "MtCO2e", "Research", "Viet Nam NDC 2022", URL_NDC, "High", "15.8% below 2030 BAU.", False),
    ("NDC conditional reduction", "Target", 403.7, "MtCO2e", "Research", "Viet Nam NDC 2022", URL_NDC, "High", "43.5% below 2030 BAU with international support.", False),
    ("Pilot annual emissions quota", "Coverage", 255.5, "MtCO2e/year", "Research", "VnExpress: 511 MtCO2e over 2025-2026", URL_ARTICLE, "Medium", "Annualized from the two-year quota; used as a pilot coverage proxy.", False),
    ("Price response exponent", "Model", 0.85, "exponent", "Model Input", "Analyst assumption", "", "Low", "Allows diminishing response as carbon price rises; change in sensitivity analysis.", True),
    ("Efficiency-to-emissions pass-through", "Model", 0.60, "%", "Model Input", "Analyst assumption", "", "Low", "Share of energy-efficiency gain assumed to translate into economy-wide GHG reduction.", True),
    ("Recycling GDP offset at threshold", "Model", 0.25, "% of carbon-price GDP drag", "Model Input", "Analyst assumption", "", "Low", "Directional offset only; not stated in the article.", True),
]

SECTORS = [
    ("Energy", 678.4, 1.20, "Power, fuel combustion, transport and industrial energy"),
    ("Agriculture", 112.1, 0.30, "Lower direct price responsiveness; non-CO2 gases dominate"),
    ("Waste", 46.3, 0.50, "Moderate response where methane capture is economic"),
    ("Industrial processes", 140.3, 0.80, "Cement, steel, chemicals and HFCs"),
]

CASES = [
    # id, label, price, coverage, exemption, recycling, efficiency, collection, leakage, feasibility, narrative
    ("pilot_ets", "Pilot ETS", 5.0, 255.5 / 927.9, 0.80, 0.20, 0.01, 0.90, 0.20, 4.0, "Narrow learning phase; high free allocation limits both burden and impact."),
    ("tax_only", "$15 Tax Only", 15.0, 1.00, 0.00, 0.00, 0.00, 0.95, 0.00, 3.0, "Reproduces the article thesis with no recycling or efficiency package; the 2.9% anchor is already treated as net of baseline leakage."),
    ("just_transition", "$15 Just Transition", 15.0, 1.00, 0.00, 0.30, 0.05, 0.95, 0.05, 4.0, "Article-consistent price plus targeted recycling and a partial efficiency push."),
    ("protected_industry", "$15 Protected Industry", 15.0, 0.70, 0.20, 0.30, 0.05, 0.90, 0.20, 3.0, "Partial coverage and exemptions reduce competitiveness risk but weaken abatement."),
    ("accelerated", "$30 Accelerated", 30.0, 1.00, 0.00, 0.40, 0.10, 0.95, 0.05, 2.5, "Higher price paired with full efficiency benchmark and stronger recycling."),
]

SCENARIOS = [
    # id, response multiplier, GDP drag multiplier, efficiency realization, probability, narrative
    ("base", 1.00, 1.00, 1.00, 0.55, "IPSS calibration holds and implementation is broadly effective."),
    ("bull", 1.25, 0.80, 1.20, 0.22, "Firms abate faster, efficiency measures deliver, and macro adjustment is smoother."),
    ("bear", 0.75, 1.25, 0.70, 0.23, "Weak response, slower efficiency delivery, and higher near-term macro friction."),
]


def safe_name(value: str) -> str:
    return "".join(ch.lower() if ch.isalnum() else "_" for ch in value).strip("_").replace("__", "_")


def add_name(wb: Workbook, name: str, sheet: str, cell: str) -> None:
    ref = f"{quote_sheetname(sheet)}!{absolute_coordinate(cell)}"
    wb.defined_names[name] = DefinedName(name, attr_text=ref)


def title(ws, text: str, end_col: int = 10, subtitle: str | None = None) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_col)
    c = ws.cell(1, 1, text)
    c.fill = PatternFill("solid", fgColor=DARK)
    c.font = Font(name="Roboto", size=16, bold=True, color=WHITE)
    c.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 30
    if subtitle:
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=end_col)
        s = ws.cell(2, 1, subtitle)
        s.font = Font(name="Roboto", size=9, italic=True, color=MID_GRAY)
        s.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[2].height = 28


def section(ws, row: int, text: str, end_col: int) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=end_col)
    c = ws.cell(row, 1, text)
    c.fill = PatternFill("solid", fgColor=GRAY)
    c.font = Font(name="Roboto", size=9, bold=True, color=WHITE)
    c.border = section_border
    c.alignment = Alignment(vertical="center")
    ws.row_dimensions[row].height = 20


def header_row(ws, row: int, columns: list[str], fills: dict[int, str] | None = None) -> None:
    for col, value in enumerate(columns, 1):
        c = ws.cell(row, col, value)
        c.fill = PatternFill("solid", fgColor=(fills or {}).get(col, LIGHT_GRAY))
        c.font = Font(name="Roboto", size=9, bold=True, color=WHITE if (fills or {}).get(col) in {DARK_GREEN, BLUE, AMBER, GRAY, DARK} else BLACK)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = Border(bottom=medium)
    ws.row_dimensions[row].height = 30


def input_style(cell, research: bool = False, key: bool = False) -> None:
    cell.fill = PatternFill("solid", fgColor=YELLOW if key else (RESEARCH_GREEN if research else INPUT_BLUE))
    cell.font = Font(name="Roboto", size=9, color=RESEARCH_FONT if research else INPUT_FONT)


def formula_style(cell) -> None:
    cell.font = Font(name="Roboto", size=9, color=BLACK)


def body_style(cell, wrap: bool = False) -> None:
    cell.font = Font(name="Roboto", size=9, color=BLACK)
    cell.alignment = Alignment(vertical="top", wrap_text=wrap)


def configure_sheet(ws, widths: dict[str, float], freeze: str = "B4", landscape: bool = True) -> None:
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = freeze
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.page_setup.orientation = "landscape" if landscape else "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.oddFooter.center.text = "Vietnam carbon tax impact simulation | v1.0 | 14 Jul 2026"
    ws.sheet_properties.tabColor = BLUE


def build_cover(wb: Workbook) -> None:
    ws = wb.create_sheet("Cover")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = DARK
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 66
    ws.merge_cells("B2:C4")
    c = ws["B2"]
    c.value = "VIETNAM CARBON TAX\nIMPACT SIMULATION"
    c.fill = PatternFill("solid", fgColor=DARK)
    c.font = Font(name="Roboto", size=22, bold=True, color=WHITE)
    c.alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 32
    ws.row_dimensions[3].height = 32
    ws.row_dimensions[4].height = 32
    ws["B6"] = "Decision question"
    ws["C6"] = "Does a USD15/tCO2e price plausibly deliver ~3% emissions reduction, and which policy package best limits the macro and equity trade-offs?"
    ws["B7"] = "Model type"
    ws["C7"] = "CGE-anchor replication + transparent policy lever simulation; not a forecast"
    ws["B8"] = "Reference periods"
    ws["C8"] = "2030 BAU emissions; 2024 GDP scale proxy; policy calibration published 13 Jul 2026"
    ws["B9"] = "Version / status"
    ws["C9"] = "v1.0 / Directional decision model"
    ws["B10"] = "Prepared"
    ws["C10"] = date(2026, 7, 14)
    ws["C10"].number_format = "dd mmm yyyy"
    ws["B12"] = "How to use"
    ws["C12"] = "Start with Executive Summary. Change blue cells in Scenario Comparison and Assumptions Log. All outputs recalculate through named ranges."
    ws["B14"] = "Confidence"
    ws["C14"] = "Directional — core article anchors are medium confidence; behavioural and recycling offsets are explicit low-confidence model assumptions."
    for r in range(6, 15):
        ws.cell(r, 2).font = Font(name="Roboto", size=9, bold=True, color=DARK_GREEN)
        ws.cell(r, 3).font = Font(name="Roboto", size=9)
        ws.cell(r, 3).alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = 28 if r in {6, 7, 8, 12, 14} else 20
    ws.print_area = "A1:C16"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1


def build_assumptions(wb: Workbook) -> dict[str, int]:
    ws = wb.create_sheet("Assumptions Log")
    title(ws, "ASSUMPTIONS LOG", 13, "Green = research anchor; blue = analyst/model input; yellow = key swing variable. All research inputs carry source URLs and comments.")
    headers = ["Assumption", "Category", "Base value", "Bull value", "Bear value", "Unit", "Source type", "Source title", "Citation URL", "Confidence", "Owner", "Date set", "Notes / caveats"]
    header_row(ws, 4, headers, {3: DARK_GREEN, 4: BLUE, 5: AMBER})
    assumption_rows: dict[str, int] = {}
    for idx, item in enumerate(ASSUMPTIONS, 5):
        name, category, value, unit, source_type, source_title, url, confidence, notes, key = item
        assumption_rows[safe_name(name)] = idx
        values = [name, category, value, value, value, unit, source_type, source_title, url, confidence, "Policy model", date(2026, 7, 14), notes]
        for col, val in enumerate(values, 1):
            c = ws.cell(idx, col, val)
            body_style(c, wrap=col in {1, 8, 9, 13})
        research = source_type == "Research"
        for col in (3, 4, 5):
            input_style(ws.cell(idx, col), research=research, key=key)
            if url:
                ws.cell(idx, col).comment = Comment(f"Source: {source_title}\n{url}\nCaveat: {notes}", "Codex")
        if isinstance(value, float) and (unit.startswith("%") or "%" in unit):
            for col in (3, 4, 5):
                ws.cell(idx, col).number_format = "0.0%"
        elif "USD" in unit:
            for col in (3, 4, 5):
                ws.cell(idx, col).number_format = '#,##0.0;[Red](#,##0.0);"-"'
        else:
            for col in (3, 4, 5):
                ws.cell(idx, col).number_format = '#,##0.0;[Red](#,##0.0);"-"'
        for scenario, col in (("base", 3), ("bull", 4), ("bear", 5)):
            add_name(wb, f"{scenario}_{safe_name(name)}", ws.title, ws.cell(idx, col).coordinate)

    scenario_start = 5 + len(ASSUMPTIONS) + 2
    section(ws, scenario_start, "SCENARIO CALIBRATION — only four assumptions vary", 13)
    header_row(ws, scenario_start + 1, ["Scenario", "Response multiplier", "GDP drag multiplier", "Efficiency realization", "Probability", "Narrative"])
    for i, (scenario_id, response, gdp_drag, eff_real, probability, narrative) in enumerate(SCENARIOS, scenario_start + 2):
        vals = [scenario_id.title(), response, gdp_drag, eff_real, probability, narrative]
        for col, val in enumerate(vals, 1):
            ws.cell(i, col, val)
            body_style(ws.cell(i, col), wrap=col == 6)
        for col in range(2, 6):
            input_style(ws.cell(i, col), research=False, key=True)
        ws.cell(i, 5).number_format = "0%"
        for suffix, col in (("response_multiplier", 2), ("gdp_drag_multiplier", 3), ("efficiency_realization", 4), ("probability", 5)):
            add_name(wb, f"scenario_{scenario_id}_{suffix}", ws.title, ws.cell(i, col).coordinate)

    widths = {"A": 34, "B": 16, "C": 13, "D": 13, "E": 13, "F": 19, "G": 14, "H": 32, "I": 42, "J": 12, "K": 15, "L": 13, "M": 52}
    configure_sheet(ws, widths, "C5")
    ws.auto_filter.ref = f"A4:M{4 + len(ASSUMPTIONS)}"
    return assumption_rows


def build_scenario_comparison(wb: Workbook) -> dict[str, int]:
    ws = wb.create_sheet("Scenario Comparison")
    title(ws, "POLICY CASES — INPUT LEVERS", 18, "Edit blue cells. Base outputs are formula-linked from the model engine; the lower table shows Base/Bull/Bear uncertainty for the featured just-transition package.")
    headers = ["Case", "Carbon price\n(USD/tCO2e)", "Coverage\n(%)", "Exemption / free allocation\n(%)", "Revenue recycling\n(%)", "Efficiency gain\n(%)", "Collection efficiency\n(%)", "Leakage\n(%)", "Feasibility\n(1-5)", "Net emissions reduction\n(%)", "Reduction\n(MtCO2e)", "Net GDP impact\n(%)", "Net GDP impact\n(USD bn)", "Net fiscal proceeds\n(USD bn)", "Bottom households compensated\n(%)", "NDC unconditional progress\n(%)", "Revenue / t abated\n(USD/t)", "Narrative"]
    header_row(ws, 4, headers)
    case_rows: dict[str, int] = {}
    for row, case in enumerate(CASES, 5):
        case_id, label, price, coverage, exemption, recycling, efficiency, collection, leakage, feasibility, narrative = case
        case_rows[case_id] = row
        inputs = [label, price, coverage, exemption, recycling, efficiency, collection, leakage, feasibility]
        for col, val in enumerate(inputs, 1):
            ws.cell(row, col, val)
            body_style(ws.cell(row, col), wrap=col == 1)
            if col >= 2:
                input_style(ws.cell(row, col))
        for col in range(3, 9):
            ws.cell(row, col).number_format = "0.0%"
        ws.cell(row, 18, narrative)
        body_style(ws.cell(row, 18), wrap=True)
        for suffix, col in (("price", 2), ("coverage", 3), ("exemption", 4), ("recycling", 5), ("efficiency", 6), ("collection", 7), ("leakage", 8), ("feasibility", 9)):
            add_name(wb, f"{case_id}_{suffix}", ws.title, ws.cell(row, col).coordinate)

    scenario_row = 12
    section(ws, scenario_row, "FEATURED CASE RANGE — $15 JUST TRANSITION", 18)
    header_row(ws, scenario_row + 1, ["Scenario", "Net emissions reduction (%)", "Reduction (MtCO2e)", "Net GDP impact (%)", "Net GDP impact (USD bn)", "Net fiscal proceeds (USD bn)", "Households compensated (%)", "NDC unconditional progress (%)", "Probability", "What would need to be true"])
    for i, (scenario_id, *_rest, narrative) in enumerate(SCENARIOS, scenario_row + 2):
        ws.cell(i, 1, scenario_id.title())
        ws.cell(i, 9, f"=scenario_{scenario_id}_probability")
        ws.cell(i, 9).number_format = "0%"
        ws.cell(i, 10, narrative)
        ws.cell(i, 10).alignment = Alignment(wrap_text=True)
        fill = {"base": DARK_GREEN, "bull": BLUE, "bear": AMBER}[scenario_id]
        ws.cell(i, 1).fill = PatternFill("solid", fgColor=fill)
        ws.cell(i, 1).font = Font(name="Roboto", size=9, bold=True, color=WHITE)

    widths = {"A": 24, "B": 14, "C": 12, "D": 18, "E": 15, "F": 14, "G": 16, "H": 12, "I": 12, "J": 17, "K": 16, "L": 15, "M": 16, "N": 17, "O": 20, "P": 18, "Q": 16, "R": 54}
    configure_sheet(ws, widths, "B5")
    ws.print_area = "A1:Q18"
    return case_rows


def build_engine(wb: Workbook, case_rows: dict[str, int]) -> dict[tuple[str, str], int]:
    ws = wb.create_sheet("Revenue Build-Up")
    title(ws, "EMISSIONS IMPACT BUILD-UP", 35, "Policy-model equivalent of a revenue build-up: each case/scenario is calculated from named input ranges. Formula columns are black; no hardcoded outputs.")
    headers = [
        "Case", "Scenario", "Price", "Coverage", "Exemption", "Recycling", "Efficiency", "Collection", "Leakage",
        "Response mult.", "GDP drag mult.", "Efficiency realization", "Effective coverage", "Price reduction", "Efficiency reduction", "Gross reduction", "Net reduction", "Reduction Mt", "Residual Mt", "Taxable Mt", "Gross revenue $bn", "Net collected $bn", "Recycled $bn", "Net budget $bn", "Carbon GDP drag", "Efficiency GDP benefit", "Recycling GDP offset", "Net GDP impact", "GDP impact $bn", "Trade impact", "Investment impact", "Govt revenue index", "HH compensation", "NDC progress", "Revenue/t abated"
    ]
    header_row(ws, 4, headers)
    engine_rows: dict[tuple[str, str], int] = {}
    row = 5
    for case in CASES:
        case_id, label = case[0], case[1]
        for scenario in SCENARIOS:
            scenario_id = scenario[0]
            engine_rows[(case_id, scenario_id)] = row
            ws.cell(row, 1, label)
            ws.cell(row, 2, scenario_id.title())
            ws.cell(row, 3, f"={case_id}_price")
            ws.cell(row, 4, f"={case_id}_coverage")
            ws.cell(row, 5, f"={case_id}_exemption")
            ws.cell(row, 6, f"={case_id}_recycling")
            ws.cell(row, 7, f"={case_id}_efficiency")
            ws.cell(row, 8, f"={case_id}_collection")
            ws.cell(row, 9, f"={case_id}_leakage")
            ws.cell(row, 10, f"=scenario_{scenario_id}_response_multiplier")
            ws.cell(row, 11, f"=scenario_{scenario_id}_gdp_drag_multiplier")
            ws.cell(row, 12, f"=scenario_{scenario_id}_efficiency_realization")
            ws.cell(row, 13, f"=D{row}*(1-E{row})")
            ws.cell(row, 14, f"=MIN(1,base_anchor_emissions_reduction*(C{row}/base_anchor_carbon_price)^base_price_response_exponent*M{row}*J{row})")
            ws.cell(row, 15, f"=MIN(1-N{row},G{row}*base_efficiency_to_emissions_pass_through*L{row})")
            ws.cell(row, 16, f"=N{row}+O{row}")
            ws.cell(row, 17, f"=P{row}*(1-I{row})")
            ws.cell(row, 18, f"=base_baseline_ghg_emissions*Q{row}")
            ws.cell(row, 19, f"=base_baseline_ghg_emissions-R{row}")
            ws.cell(row, 20, f"=base_baseline_ghg_emissions*M{row}*(1-P{row})")
            ws.cell(row, 21, f"=C{row}*T{row}/1000")
            ws.cell(row, 22, f"=U{row}*H{row}")
            ws.cell(row, 23, f"=V{row}*F{row}")
            ws.cell(row, 24, f"=V{row}-W{row}")
            ws.cell(row, 25, f"=-base_anchor_gdp_drag*(C{row}/base_anchor_carbon_price)^base_price_response_exponent*M{row}*K{row}")
            ws.cell(row, 26, f"=IF(base_efficiency_benchmark=0,0,G{row}/base_efficiency_benchmark*base_efficiency_gdp_benefit_at_benchmark*L{row})")
            ws.cell(row, 27, f"=-Y{row}*MIN(1,F{row}/base_revenue_recycling_threshold)*base_recycling_gdp_offset_at_threshold")
            ws.cell(row, 28, f"=Y{row}+Z{row}+AA{row}")
            ws.cell(row, 29, f"=base_reference_nominal_gdp*AB{row}")
            ws.cell(row, 30, f"=-base_anchor_trade_drag*(C{row}/base_anchor_carbon_price)^base_price_response_exponent*M{row}*K{row}")
            ws.cell(row, 31, f"=-base_anchor_investment_drag*(C{row}/base_anchor_carbon_price)^base_price_response_exponent*M{row}*K{row}")
            ws.cell(row, 32, f"=base_anchor_government_revenue_index_uplift*(C{row}/base_anchor_carbon_price)^base_price_response_exponent*M{row}*H{row}")
            ws.cell(row, 33, f"=MIN(base_bottom_households_potentially_compensated,base_bottom_households_potentially_compensated*F{row}/base_revenue_recycling_threshold)")
            ws.cell(row, 34, f"=R{row}/base_ndc_unconditional_reduction")
            ws.cell(row, 35, f"=IF(R{row}=0,0,V{row}*1000/R{row})")
            for col in range(1, 36):
                body_style(ws.cell(row, col))
            for col in [4, 5, 6, 7, 8, 9, 13, 14, 15, 16, 17, 25, 26, 27, 28, 30, 31, 32, 33, 34]:
                ws.cell(row, col).number_format = "0.0%"
            for col in [18, 19, 20, 21, 22, 23, 24, 29, 35]:
                ws.cell(row, col).number_format = '#,##0.0;[Red](#,##0.0);"-"'
            output_names = {
                "emissions_reduction_pct": 17, "emissions_reduction_mt": 18, "residual_emissions_mt": 19,
                "gross_fiscal_revenue_bn": 21, "net_fiscal_revenue_bn": 22, "recycled_revenue_bn": 23,
                "net_budget_revenue_bn": 24, "net_gdp_impact_pct": 28, "net_gdp_impact_bn": 29,
                "trade_impact_pct": 30, "investment_impact_pct": 31, "government_revenue_index_pct": 32,
                "household_compensation_pct": 33, "ndc_progress_pct": 34, "fiscal_revenue_per_t_abated": 35,
            }
            for suffix, col in output_names.items():
                add_name(wb, f"{case_id}_{scenario_id}_{suffix}", ws.title, ws.cell(row, col).coordinate)
            row += 1

    # Link base outputs and featured scenario range back to Scenario Comparison.
    sc = wb["Scenario Comparison"]
    for case_id, case_row in case_rows.items():
        for col, suffix in ((10, "emissions_reduction_pct"), (11, "emissions_reduction_mt"), (12, "net_gdp_impact_pct"), (13, "net_gdp_impact_bn"), (14, "net_fiscal_revenue_bn"), (15, "household_compensation_pct"), (16, "ndc_progress_pct"), (17, "fiscal_revenue_per_t_abated")):
            sc.cell(case_row, col, f"={case_id}_base_{suffix}")
            formula_style(sc.cell(case_row, col))
        for col in [10, 12, 15, 16]:
            sc.cell(case_row, col).number_format = "0.0%"
        for col in [11, 13, 14, 17]:
            sc.cell(case_row, col).number_format = '#,##0.0;[Red](#,##0.0);"-"'
    for idx, scenario_id in enumerate([s[0] for s in SCENARIOS], 14):
        for col, suffix in ((2, "emissions_reduction_pct"), (3, "emissions_reduction_mt"), (4, "net_gdp_impact_pct"), (5, "net_gdp_impact_bn"), (6, "net_fiscal_revenue_bn"), (7, "household_compensation_pct"), (8, "ndc_progress_pct")):
            sc.cell(idx, col, f"=just_transition_{scenario_id}_{suffix}")
        for col in [2, 4, 7, 8]:
            sc.cell(idx, col).number_format = "0.0%"
        for col in [3, 5, 6]:
            sc.cell(idx, col).number_format = '#,##0.0;[Red](#,##0.0);"-"'

    widths = {get_column_letter(i): (24 if i == 1 else 14) for i in range(1, 36)}
    configure_sheet(ws, widths, "C5")
    ws.auto_filter.ref = f"A4:AI{row - 1}"
    return engine_rows


def build_executive_summary(wb: Workbook) -> None:
    ws = wb.create_sheet("Executive Summary")
    title(ws, "USD15/tCO2e CAN REPLICATE THE ~3% THESIS — BUT POLICY DESIGN DETERMINES WHETHER THE GDP TRADE-OFF PERSISTS", 12, "Executive readout: thesis replication, output ranges, and recommended package. Values update when blue inputs change.")
    section(ws, 4, "KEY FINDINGS", 12)
    findings = [
        '="1. The $15 tax-only case produces "&TEXT(tax_only_base_emissions_reduction_pct,"0.0%")&" net reduction ("&TEXT(tax_only_base_emissions_reduction_mt,"0.0")&" MtCO2e), closely testing the article\'s 2.9% anchor."',
        '="2. Adding 5% efficiency and 30% recycling changes the modeled GDP impact from "&TEXT(tax_only_base_net_gdp_impact_pct,"0.0%")&" to "&TEXT(just_transition_base_net_gdp_impact_pct,"0.0%")&" while increasing emissions reduction to "&TEXT(just_transition_base_emissions_reduction_pct,"0.0%")&"."',
        '="3. Across five Base policy cases, net emissions reduction ranges from "&TEXT(MIN(pilot_ets_base_emissions_reduction_pct,tax_only_base_emissions_reduction_pct,just_transition_base_emissions_reduction_pct,protected_industry_base_emissions_reduction_pct,accelerated_base_emissions_reduction_pct),"0.0%")&" to "&TEXT(MAX(pilot_ets_base_emissions_reduction_pct,tax_only_base_emissions_reduction_pct,just_transition_base_emissions_reduction_pct,protected_industry_base_emissions_reduction_pct,accelerated_base_emissions_reduction_pct),"0.0%")&"."',
        '="4. A 30% recycling share directs "&TEXT(just_transition_base_recycled_revenue_bn,"$0.0")&"bn to transition support and reaches up to "&TEXT(just_transition_base_household_compensation_pct,"0%")&" of households under the article\'s compensation anchor."',
        '="5. Even the $30 accelerated case delivers only "&TEXT(accelerated_base_ndc_progress_pct,"0%")&" of the unconditional NDC reduction in this directional model; carbon pricing is a package component, not a standalone solution."',
    ]
    for i, formula in enumerate(findings, 5):
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=12)
        ws.cell(i, 1, formula)
        ws.cell(i, 1).font = Font(name="Roboto", size=10, color=BLACK)
        ws.cell(i, 1).alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[i].height = 30

    section(ws, 11, "POLICY CASE COMPARISON — BASE SCENARIO", 12)
    headers = ["Metric", "Pilot ETS", "$15 Tax Only", "$15 Just Transition", "$15 Protected", "$30 Accelerated"]
    header_row(ws, 12, headers, {2: GRAY, 3: AMBER, 4: DARK_GREEN, 5: BLUE, 6: BLUE})
    metrics = [
        ("Net emissions reduction (%)", "emissions_reduction_pct", "0.0%"),
        ("Reduction (MtCO2e)", "emissions_reduction_mt", '#,##0.0;[Red](#,##0.0);"-"'),
        ("Net GDP impact (%)", "net_gdp_impact_pct", "0.0%"),
        ("Net GDP impact (USD bn)", "net_gdp_impact_bn", '#,##0.0;[Red](#,##0.0);"-"'),
        ("Net fiscal proceeds (USD bn)", "net_fiscal_revenue_bn", '#,##0.0;[Red](#,##0.0);"-"'),
        ("Households compensated (%)", "household_compensation_pct", "0%"),
        ("NDC unconditional progress (%)", "ndc_progress_pct", "0%"),
    ]
    case_ids = [c[0] for c in CASES]
    for r, (label, suffix, fmt) in enumerate(metrics, 13):
        ws.cell(r, 1, label)
        ws.cell(r, 1).font = Font(name="Roboto", size=9, bold=True)
        for c, case_id in enumerate(case_ids, 2):
            ws.cell(r, c, f"={case_id}_base_{suffix}")
            ws.cell(r, c).number_format = fmt
            formula_style(ws.cell(r, c))

    section(ws, 22, "RECOMMENDATION", 12)
    ws.merge_cells("A23:L24")
    ws["A23"] = "Proceed with a staged $15/tCO2e broad price only as a package: preserve the marginal price signal, recycle at least 30% of proceeds to lower-income households, and pair it with measurable energy-efficiency delivery. Treat the tax-only case as the counterfactual—not the recommended design."
    ws["A23"].font = Font(name="Roboto", size=11, bold=True, color=DARK_GREEN)
    ws["A23"].alignment = Alignment(wrap_text=True, vertical="center")
    ws["A25"] = "Confidence"
    ws["B25"] = "Directional / Medium-Low"
    ws["D25"] = "Critical uncertainty"
    ws["E25"] = "Behavioural response and the macro benefit of recycling are not identified in the article."
    ws.merge_cells("E25:L25")

    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "Base-case emissions reduction by policy package"
    chart.y_axis.title = "Reduction (%)"
    chart.height = 4.8
    chart.width = 11.5
    data = Reference(ws, min_col=2, max_col=6, min_row=13, max_row=13)
    cats = Reference(ws, min_col=2, max_col=6, min_row=12, max_row=12)
    chart.add_data(data, from_rows=True, titles_from_data=False)
    chart.set_categories(cats)
    chart.legend = None
    ws.add_chart(chart, "G12")

    widths = {"A": 32, "B": 16, "C": 16, "D": 18, "E": 18, "F": 17, "G": 3, "H": 13, "I": 13, "J": 13, "K": 13, "L": 13}
    configure_sheet(ws, widths, "A4")
    ws.sheet_properties.tabColor = DARK_GREEN
    ws.print_area = "A1:L26"


def build_decision_matrix(wb: Workbook) -> None:
    ws = wb.create_sheet("Decision Matrix")
    title(ws, "$30 ACCELERATED MAXIMIZES MODELLED IMPACT; $15 JUST TRANSITION IS THE PRAGMATIC STARTING PACKAGE", 12, "Scores are formula-driven except feasibility, which is an editable policy-judgment input on Scenario Comparison. Highest score is not automatically the recommended launch sequence.")
    headers = ["Criterion", "Weight", "Pilot ETS", "$15 Tax Only", "$15 Just Transition", "$15 Protected", "$30 Accelerated", "Scoring logic"]
    header_row(ws, 4, headers)
    criteria = [
        ("Emissions impact", 0.30, "emissions", "Higher reduction scores better"),
        ("Macro protection", 0.25, "gdp", "Less negative / more positive GDP impact scores better"),
        ("Equity support", 0.20, "equity", "Higher household compensation coverage scores better"),
        ("Fiscal capacity", 0.10, "fiscal", "Higher net fiscal proceeds scores better"),
        ("Implementation feasibility", 0.15, "feasibility", "Manual 1-5 policy judgment"),
    ]
    case_ids = [c[0] for c in CASES]
    for r, (criterion, weight, metric, logic) in enumerate(criteria, 5):
        ws.cell(r, 1, criterion)
        ws.cell(r, 2, weight)
        ws.cell(r, 2).number_format = "0%"
        input_style(ws.cell(r, 2))
        ws.cell(r, 8, logic)
        for c, case_id in enumerate(case_ids, 3):
            if metric == "emissions":
                formula = f"=1+4*{case_id}_base_emissions_reduction_pct/MAX(pilot_ets_base_emissions_reduction_pct,tax_only_base_emissions_reduction_pct,just_transition_base_emissions_reduction_pct,protected_industry_base_emissions_reduction_pct,accelerated_base_emissions_reduction_pct)"
            elif metric == "gdp":
                formula = f"=1+4*({case_id}_base_net_gdp_impact_pct-MIN(pilot_ets_base_net_gdp_impact_pct,tax_only_base_net_gdp_impact_pct,just_transition_base_net_gdp_impact_pct,protected_industry_base_net_gdp_impact_pct,accelerated_base_net_gdp_impact_pct))/(MAX(pilot_ets_base_net_gdp_impact_pct,tax_only_base_net_gdp_impact_pct,just_transition_base_net_gdp_impact_pct,protected_industry_base_net_gdp_impact_pct,accelerated_base_net_gdp_impact_pct)-MIN(pilot_ets_base_net_gdp_impact_pct,tax_only_base_net_gdp_impact_pct,just_transition_base_net_gdp_impact_pct,protected_industry_base_net_gdp_impact_pct,accelerated_base_net_gdp_impact_pct))"
            elif metric == "equity":
                formula = f"=1+4*{case_id}_base_household_compensation_pct/base_bottom_households_potentially_compensated"
            elif metric == "fiscal":
                formula = f"=1+4*{case_id}_base_net_fiscal_revenue_bn/MAX(pilot_ets_base_net_fiscal_revenue_bn,tax_only_base_net_fiscal_revenue_bn,just_transition_base_net_fiscal_revenue_bn,protected_industry_base_net_fiscal_revenue_bn,accelerated_base_net_fiscal_revenue_bn)"
            else:
                formula = f"={case_id}_feasibility"
            ws.cell(r, c, formula)
            ws.cell(r, c).number_format = "0.0x"
    total_row = 11
    ws.cell(total_row, 1, "WEIGHTED SCORE")
    ws.cell(total_row, 2, "=SUM(B5:B9)")
    ws.cell(total_row, 2).number_format = "0%"
    for c in range(3, 8):
        letter = get_column_letter(c)
        ws.cell(total_row, c, f"=SUMPRODUCT($B$5:$B$9,{letter}$5:{letter}$9)")
        ws.cell(total_row, c).number_format = "0.00x"
        ws.cell(total_row, c).border = total_border
        ws.cell(total_row, c).font = Font(name="Roboto", bold=True)
    ws.cell(13, 1, "Highest model score")
    ws.cell(13, 2, "=INDEX(C4:G4,1,MATCH(MAX(C11:G11),C11:G11,0))")
    ws.cell(13, 2).fill = PatternFill("solid", fgColor=RESEARCH_GREEN)
    ws.cell(13, 2).font = Font(name="Roboto", bold=True, color=DARK_GREEN)
    ws.merge_cells("B13:D13")
    widths = {"A": 28, "B": 12, "C": 15, "D": 16, "E": 19, "F": 17, "G": 17, "H": 48}
    configure_sheet(ws, widths, "C5")
    ws.sheet_properties.tabColor = DARK_GREEN
    ws.print_area = "A1:H14"


def build_fiscal_social(wb: Workbook) -> None:
    ws = wb.create_sheet("Cost & Investment")
    title(ws, "FISCAL PROCEEDS CREATE ROOM FOR COMPENSATION — BUT GROSS TAX COLLECTION IS NOT AN ECONOMIC BENEFIT", 10, "Policy-model equivalent of cost/investment build-up. Gross proceeds, recycling, and retained budget are shown separately.")
    headers = ["Metric", "Pilot ETS", "$15 Tax Only", "$15 Just Transition", "$15 Protected", "$30 Accelerated", "Interpretation"]
    header_row(ws, 4, headers)
    rows = [
        ("Gross fiscal revenue (USD bn)", "gross_fiscal_revenue_bn", "Mechanical carbon charge on the post-abatement taxable base."),
        ("Net collected revenue (USD bn)", "net_fiscal_revenue_bn", "After collection efficiency."),
        ("Recycled to households/transition (USD bn)", "recycled_revenue_bn", "Earmarked support; not retained budget."),
        ("Retained net budget (USD bn)", "net_budget_revenue_bn", "Net collected less recycling."),
        ("Households potentially compensated (%)", "household_compensation_pct", "Anchored to article: 30% recycling can compensate bottom 50%."),
        ("Fiscal revenue per tonne abated (USD/t)", "fiscal_revenue_per_t_abated", "A fiscal-yield metric, not marginal abatement cost."),
    ]
    case_ids = [c[0] for c in CASES]
    for r, (label, suffix, note) in enumerate(rows, 5):
        ws.cell(r, 1, label)
        ws.cell(r, 7, note)
        ws.cell(r, 7).alignment = Alignment(wrap_text=True)
        for c, case_id in enumerate(case_ids, 2):
            ws.cell(r, c, f"={case_id}_base_{suffix}")
            ws.cell(r, c).number_format = "0.0%" if "pct" in suffix else '#,##0.0;[Red](#,##0.0);"-"'
    section(ws, 13, "EQUITY DESIGN TEST", 10)
    ws["A14"] = "Recycling share required to reach bottom 50%"
    ws["B14"] = "=base_revenue_recycling_threshold"
    ws["B14"].number_format = "0%"
    ws["A15"] = "Just-transition recycling share"
    ws["B15"] = "=just_transition_recycling"
    ws["B15"].number_format = "0%"
    ws["A16"] = "Threshold met?"
    ws["B16"] = '=IF(B15>=B14,"YES","NO")'
    widths = {"A": 42, "B": 16, "C": 16, "D": 18, "E": 17, "F": 17, "G": 54}
    configure_sheet(ws, widths, "B5")


def build_macro(wb: Workbook) -> None:
    ws = wb.create_sheet("P&L Waterfall")
    title(ws, "EFFICIENCY IS THE PRIMARY MODELLED BRIDGE FROM A CARBON-PRICE GDP DRAG TO A GROWTH-NEUTRAL PACKAGE", 11, "Policy-model equivalent of a P&L waterfall. GDP effects are additive directional proxies, not a full general-equilibrium forecast.")
    headers = ["Macro component", "Pilot ETS", "$15 Tax Only", "$15 Just Transition", "$15 Protected", "$30 Accelerated", "Source / treatment"]
    header_row(ws, 4, headers)
    rows = [
        ("Carbon-price GDP drag (%)", "carbon", "IPSS anchor scaled by price, effective coverage, and scenario drag multiplier."),
        ("Efficiency GDP benefit (%)", "eff", "IPSS anchor: 10% efficiency -> +1% GDP, scaled linearly."),
        ("Recycling GDP offset (%)", "recycle", "Low-confidence model assumption: partial offset of carbon-price drag."),
        ("NET GDP IMPACT (%)", "net_gdp_impact_pct", "Sum of the three components."),
        ("Net GDP impact (USD bn)", "net_gdp_impact_bn", "Applied to 2024 GDP only as an indicative scale proxy."),
        ("Trade impact (%)", "trade_impact_pct", "IPSS -0.9% anchor scaled with price and effective coverage."),
        ("Investment impact (%)", "investment_impact_pct", "IPSS -1.3% anchor scaled with price and effective coverage."),
        ("Government revenue index impact (%)", "government_revenue_index_pct", "Separate from mechanical tax proceeds; follows IPSS +0.41% anchor."),
    ]
    case_ids = [c[0] for c in CASES]
    engine = wb["Revenue Build-Up"]
    # Add named outputs for macro components that were not exposed earlier.
    for case_id in case_ids:
        row = next(r for (cid, sid), r in ENGINE_ROWS.items() if cid == case_id and sid == "base")
        for suffix, col in (("carbon_gdp_drag_pct", 25), ("efficiency_gdp_benefit_pct", 26), ("recycling_gdp_offset_pct", 27)):
            add_name(wb, f"{case_id}_base_{suffix}", engine.title, engine.cell(row, col).coordinate)
    for r, (label, suffix, note) in enumerate(rows, 5):
        ws.cell(r, 1, label)
        ws.cell(r, 7, note)
        ws.cell(r, 7).alignment = Alignment(wrap_text=True)
        if suffix == "carbon": actual = "carbon_gdp_drag_pct"
        elif suffix == "eff": actual = "efficiency_gdp_benefit_pct"
        elif suffix == "recycle": actual = "recycling_gdp_offset_pct"
        else: actual = suffix
        for c, case_id in enumerate(case_ids, 2):
            ws.cell(r, c, f"={case_id}_base_{actual}")
            ws.cell(r, c).number_format = "0.0%" if "pct" in actual else '#,##0.0;[Red](#,##0.0);"-"'
        if "NET" in label:
            for c in range(1, 7):
                ws.cell(r, c).border = total_border
                ws.cell(r, c).font = Font(name="Roboto", bold=True)
    widths = {"A": 38, "B": 16, "C": 16, "D": 18, "E": 17, "F": 17, "G": 60}
    configure_sheet(ws, widths, "B5")


def build_roi_ndc(wb: Workbook) -> None:
    ws = wb.create_sheet("ROI & Payback")
    title(ws, "CARBON PRICING CLOSES ONLY PART OF THE NDC GAP — POLICY YIELD, NOT FINANCIAL ROI, IS THE RELEVANT TEST", 10, "Policy-model equivalent of ROI/payback. It compares abatement and fiscal yield against Viet Nam's 2030 NDC reductions.")
    headers = ["Metric", "Pilot ETS", "$15 Tax Only", "$15 Just Transition", "$15 Protected", "$30 Accelerated", "Benchmark / implication"]
    header_row(ws, 4, headers)
    rows = [
        ("Reduction (MtCO2e)", "emissions_reduction_mt", "Absolute abatement from 2030 BAU baseline."),
        ("Unconditional NDC progress (%)", "ndc_progress_pct", "Share of 146.3 MtCO2e unconditional reduction."),
        ("Conditional NDC progress (%)", "conditional", "Share of 403.7 MtCO2e conditional reduction."),
        ("Net fiscal proceeds (USD bn)", "net_fiscal_revenue_bn", "Potential fiscal capacity before recycling."),
        ("Fiscal revenue / tonne abated (USD/t)", "fiscal_revenue_per_t_abated", "Not an abatement cost; shows tax-base-to-abatement relationship."),
    ]
    case_ids = [c[0] for c in CASES]
    for r, (label, suffix, note) in enumerate(rows, 5):
        ws.cell(r, 1, label)
        ws.cell(r, 7, note)
        ws.cell(r, 7).alignment = Alignment(wrap_text=True)
        for c, case_id in enumerate(case_ids, 2):
            if suffix == "conditional":
                formula = f"={case_id}_base_emissions_reduction_mt/base_ndc_conditional_reduction"
                fmt = "0%"
            else:
                formula = f"={case_id}_base_{suffix}"
                fmt = "0%" if "pct" in suffix else '#,##0.0;[Red](#,##0.0);"-"'
            ws.cell(r, c, formula)
            ws.cell(r, c).number_format = fmt
    section(ws, 12, "NDC BENCHMARKS", 10)
    ws["A13"] = "2030 BAU emissions (MtCO2e)"
    ws["B13"] = "=base_baseline_ghg_emissions"
    ws["A14"] = "Unconditional reduction target (MtCO2e)"
    ws["B14"] = "=base_ndc_unconditional_reduction"
    ws["A15"] = "Conditional reduction target (MtCO2e)"
    ws["B15"] = "=base_ndc_conditional_reduction"
    widths = {"A": 42, "B": 16, "C": 16, "D": 18, "E": 17, "F": 17, "G": 58}
    configure_sheet(ws, widths, "B5")


def build_sector_incidence(wb: Workbook) -> None:
    ws = wb.create_sheet("Sector Incidence")
    title(ws, "ENERGY AND INDUSTRIAL PROCESSES CARRY MOST OF THE PRICE SIGNAL AND ABATEMENT BURDEN", 10, "Sector allocation uses NDC 2030 BAU values and explicit response weights. It is an incidence illustration, not a sector CGE model.")
    headers = ["Sector", "2030 BAU (MtCO2e)", "Share of positive emissions (%)", "Response weight (x)", "Weighted response share (%)", "Just-transition reduction (MtCO2e)", "Residual emissions (MtCO2e)", "Indicative gross charge (USD bn)", "Incidence note"]
    header_row(ws, 4, headers)
    total_positive = sum(v for _, v, _, _ in SECTORS)
    for r, (sector_name, emissions, weight, note) in enumerate(SECTORS, 5):
        ws.cell(r, 1, sector_name)
        ws.cell(r, 2, emissions)
        ws.cell(r, 3, f"=B{r}/SUM($B$5:$B$8)")
        ws.cell(r, 4, weight)
        input_style(ws.cell(r, 4))
        ws.cell(r, 5, f"=B{r}*D{r}/SUMPRODUCT($B$5:$B$8,$D$5:$D$8)")
        ws.cell(r, 6, f"=just_transition_base_emissions_reduction_mt*E{r}")
        ws.cell(r, 7, f"=B{r}-F{r}")
        ws.cell(r, 8, f"=G{r}*just_transition_price*just_transition_coverage*(1-just_transition_exemption)/1000")
        ws.cell(r, 9, note)
        ws.cell(r, 9).alignment = Alignment(wrap_text=True)
        ws.cell(r, 3).number_format = "0.0%"
        ws.cell(r, 5).number_format = "0.0%"
        for c in [2, 4, 6, 7, 8]:
            ws.cell(r, c).number_format = '#,##0.0;[Red](#,##0.0);"-"'
        ws.cell(r, 2).comment = Comment(f"Source: Viet Nam NDC 2022\n{URL_NDC}", "Codex")
    ws.cell(10, 1, "TOTAL")
    for col in [2, 3, 5, 6, 7, 8]:
        letter = get_column_letter(col)
        ws.cell(10, col, f"=SUM({letter}5:{letter}8)")
        ws.cell(10, col).border = total_border
        ws.cell(10, col).font = Font(name="Roboto", bold=True)
    ws["A12"] = "LULUCF note"
    ws["B12"] = "The NDC includes a -49.2 MtCO2e land-use sink in the 927.9 MtCO2e total. The incidence table excludes that negative sink from the taxable sector allocation."
    ws.merge_cells("B12:I12")
    ws["B12"].alignment = Alignment(wrap_text=True)
    widths = {"A": 24, "B": 18, "C": 18, "D": 17, "E": 19, "F": 20, "G": 20, "H": 22, "I": 52}
    configure_sheet(ws, widths, "B5")


def build_sensitivity(wb: Workbook) -> None:
    ws = wb.create_sheet("Sensitivity Analysis")
    title(ws, "PRICE ALONE HAS DIMINISHING RETURNS; COVERAGE, EFFICIENCY, AND RESPONSE DETERMINE THE OUTPUT RANGE", 14, "All tables are formula-based (no unsupported Excel data-table feature). Edit the central assumptions in Assumptions Log or case levers in Scenario Comparison.")
    section(ws, 4, "ONE-WAY PRICE SENSITIVITY — JUST-TRANSITION DESIGN", 14)
    header_row(ws, 5, ["Carbon price (USD/tCO2e)", "Net emissions reduction (%)", "Reduction (MtCO2e)", "Net GDP impact (%)", "Net fiscal proceeds (USD bn)"])
    prices = [0, 5, 10, 15, 20, 30, 50]
    for r, price in enumerate(prices, 6):
        ws.cell(r, 1, price)
        input_style(ws.cell(r, 1))
        ws.cell(r, 2, f"=(MIN(1,base_anchor_emissions_reduction*(A{r}/base_anchor_carbon_price)^base_price_response_exponent*just_transition_coverage*(1-just_transition_exemption)*scenario_base_response_multiplier)+just_transition_efficiency*base_efficiency_to_emissions_pass_through*scenario_base_efficiency_realization)*(1-just_transition_leakage)")
        ws.cell(r, 3, f"=base_baseline_ghg_emissions*B{r}")
        ws.cell(r, 4, f"=-base_anchor_gdp_drag*(A{r}/base_anchor_carbon_price)^base_price_response_exponent*just_transition_coverage*(1-just_transition_exemption)*scenario_base_gdp_drag_multiplier+(just_transition_efficiency/base_efficiency_benchmark*base_efficiency_gdp_benefit_at_benchmark*scenario_base_efficiency_realization)+base_anchor_gdp_drag*(A{r}/base_anchor_carbon_price)^base_price_response_exponent*just_transition_coverage*(1-just_transition_exemption)*scenario_base_gdp_drag_multiplier*MIN(1,just_transition_recycling/base_revenue_recycling_threshold)*base_recycling_gdp_offset_at_threshold")
        ws.cell(r, 5, f"=A{r}*base_baseline_ghg_emissions*just_transition_coverage*(1-just_transition_exemption)*(1-B{r})/1000*just_transition_collection")
        ws.cell(r, 2).number_format = "0.0%"
        ws.cell(r, 4).number_format = "0.0%"
        for c in [3, 5]: ws.cell(r, c).number_format = '#,##0.0;[Red](#,##0.0);"-"'

    section(ws, 15, "TWO-WAY SENSITIVITY — NET EMISSIONS REDUCTION (%)", 14)
    coverages = [0.30, 0.50, 0.70, 0.85, 1.00]
    two_prices = [5, 10, 15, 20, 30, 50]
    ws.cell(16, 1, "Price / coverage")
    for c, cov in enumerate(coverages, 2):
        ws.cell(16, c, cov)
        ws.cell(16, c).number_format = "0%"
        input_style(ws.cell(16, c))
    for r, price in enumerate(two_prices, 17):
        ws.cell(r, 1, price)
        input_style(ws.cell(r, 1))
        for c in range(2, 2 + len(coverages)):
            ws.cell(r, c, f"=(MIN(1,base_anchor_emissions_reduction*($A{r}/base_anchor_carbon_price)^base_price_response_exponent*{get_column_letter(c)}$16*(1-just_transition_exemption)*scenario_base_response_multiplier)+just_transition_efficiency*base_efficiency_to_emissions_pass_through)*(1-just_transition_leakage)")
            ws.cell(r, c).number_format = "0.0%"
    ws.conditional_formatting.add(f"B17:F{16+len(two_prices)}", ColorScaleRule(start_type="min", start_color="F8696B", mid_type="percentile", mid_value=50, mid_color="FFEB84", end_type="max", end_color="63BE7B"))

    section(ws, 25, "TORNADO INPUTS — JUST-TRANSITION BASE CASE ±20%", 14)
    header_row(ws, 26, ["Driver", "Low input", "Base input", "High input", "Low emissions reduction (%)", "Base emissions reduction (%)", "High emissions reduction (%)", "Low GDP impact (%)", "Base GDP impact (%)", "High GDP impact (%)"])
    drivers = [
        ("Carbon price", "just_transition_price", 0.8, 1.2),
        ("Coverage", "just_transition_coverage", 0.8, 1.0),
        ("Response multiplier", "scenario_base_response_multiplier", 0.8, 1.2),
        ("Leakage", "just_transition_leakage", 0.8, 1.2),
        ("Efficiency gain", "just_transition_efficiency", 0.8, 1.2),
        ("GDP drag multiplier", "scenario_base_gdp_drag_multiplier", 0.8, 1.2),
    ]
    for r, (driver, name, low_mult, high_mult) in enumerate(drivers, 27):
        ws.cell(r, 1, driver)
        ws.cell(r, 2, f"={name}*{low_mult}")
        ws.cell(r, 3, f"={name}")
        ws.cell(r, 4, f"={name}*{high_mult}")
        # Explicit formulas by driver keep the audit trail legible.
        for scenario_col, input_col in ((5, "B"), (6, "C"), (7, "D")):
            price = f"{input_col}{r}" if driver == "Carbon price" else "just_transition_price"
            coverage = f"{input_col}{r}" if driver == "Coverage" else "just_transition_coverage"
            response = f"{input_col}{r}" if driver == "Response multiplier" else "scenario_base_response_multiplier"
            leakage = f"{input_col}{r}" if driver == "Leakage" else "just_transition_leakage"
            efficiency = f"{input_col}{r}" if driver == "Efficiency gain" else "just_transition_efficiency"
            ws.cell(r, scenario_col, f"=(MIN(1,base_anchor_emissions_reduction*({price}/base_anchor_carbon_price)^base_price_response_exponent*{coverage}*(1-just_transition_exemption)*{response})+{efficiency}*base_efficiency_to_emissions_pass_through)* (1-{leakage})")
            ws.cell(r, scenario_col).number_format = "0.0%"
        for scenario_col, input_col in ((8, "B"), (9, "C"), (10, "D")):
            price = f"{input_col}{r}" if driver == "Carbon price" else "just_transition_price"
            coverage = f"{input_col}{r}" if driver == "Coverage" else "just_transition_coverage"
            efficiency = f"{input_col}{r}" if driver == "Efficiency gain" else "just_transition_efficiency"
            drag = f"{input_col}{r}" if driver == "GDP drag multiplier" else "scenario_base_gdp_drag_multiplier"
            ws.cell(r, scenario_col, f"=-base_anchor_gdp_drag*({price}/base_anchor_carbon_price)^base_price_response_exponent*{coverage}*(1-just_transition_exemption)*{drag}+({efficiency}/base_efficiency_benchmark*base_efficiency_gdp_benefit_at_benchmark*scenario_base_efficiency_realization)+base_anchor_gdp_drag*({price}/base_anchor_carbon_price)^base_price_response_exponent*{coverage}*(1-just_transition_exemption)*{drag}*MIN(1,just_transition_recycling/base_revenue_recycling_threshold)*base_recycling_gdp_offset_at_threshold")
            ws.cell(r, scenario_col).number_format = "0.0%"

    line = LineChart()
    line.title = "Price sensitivity: emissions reduction and GDP impact"
    line.y_axis.title = "Impact (%)"
    line.x_axis.title = "Carbon price (USD/tCO2e)"
    line.height = 7
    line.width = 14
    data = Reference(ws, min_col=2, max_col=4, min_row=5, max_row=12)
    cats = Reference(ws, min_col=1, min_row=6, max_row=12)
    line.add_data(data, titles_from_data=True)
    line.set_categories(cats)
    ws.add_chart(line, "H5")

    widths = {"A": 25, "B": 15, "C": 15, "D": 15, "E": 18, "F": 18, "G": 18, "H": 16, "I": 16, "J": 16, "K": 13, "L": 13, "M": 13, "N": 13}
    configure_sheet(ws, widths, "B6")


def build_sources(wb: Workbook) -> None:
    ws = wb.create_sheet("Data Sources")
    title(ws, "DATA SOURCES & CITATION REGISTRY", 8, "Primary sources are preferred. The VnExpress article is used for the newly reported IPSS-UNDP calibration because the underlying study was not publicly indexed at build time.")
    headers = ["ID", "Source", "Publisher", "Date / period", "URL", "Model use", "Confidence", "Caveat"]
    header_row(ws, 4, headers)
    sources = [
        (1, "‘Áp thuế 15 USD mỗi tấn CO2…’", "VnExpress / IPSS-UNDP study summary", "13 Jul 2026", URL_ARTICLE, "$15 calibration; macro impacts; recycling and efficiency anchors", "Medium", "Secondary report of a newly released study; underlying model tables unavailable."),
        (2, "Viet Nam NDC 2022", "Government of Viet Nam / UNFCCC", "2022; 2030 BAU", URL_NDC, "BAU emissions, sector values, NDC reductions", "High", "BAU projection, not observed emissions."),
        (3, "Viet Nam country data", "World Bank", "2024", URL_WB_GDP, "Reference nominal GDP", "High", "Used only as an indicative USD scale proxy."),
        (4, "Viet Nam 2045: Growing Greener", "World Bank", "2025", URL_WB_2045, "Policy triangulation and carbon-pricing context", "High", "Not used to overwrite the article's IPSS calibration."),
        (5, "Estimating CO2 emission and revenue effects of carbon pricing", "OECD", "2022", URL_OECD_EFFECT, "External reasonableness check", "High", "Cross-country long-term estimate; not directly transferable to Viet Nam."),
    ]
    for r, source in enumerate(sources, 5):
        for c, val in enumerate(source, 1):
            ws.cell(r, c, val)
            body_style(ws.cell(r, c), wrap=c in {2, 5, 6, 8})
        ws.cell(r, 5).hyperlink = source[4]
        ws.cell(r, 5).style = "Hyperlink"
    widths = {"A": 8, "B": 38, "C": 30, "D": 16, "E": 58, "F": 42, "G": 13, "H": 52}
    configure_sheet(ws, widths, "A5")


def build_methodology(wb: Workbook) -> None:
    ws = wb.create_sheet("Methodology Notes")
    title(ws, "METHODOLOGY, LIMITATIONS, AND VERSION HISTORY", 10, "This workbook is designed to test a thesis transparently—not to reproduce the unpublished IPSS CGE model.")
    sections = [
        (4, "1. CORE EQUATIONS", [
            "Price-driven reduction = 2.9% × (price / $15)^price exponent × effective coverage × response multiplier.",
            "Efficiency reduction = efficiency gain × efficiency-to-emissions pass-through × scenario realization.",
            "Net reduction = (price reduction + efficiency reduction) × (1 − leakage).",
            "Net GDP impact = carbon-price drag + efficiency benefit + recycling offset.",
            "Tax proceeds = carbon price × post-abatement taxable emissions × collection efficiency.",
        ]),
        (12, "2. INTERPRETATION", [
            "The $15 tax-only/Base case is the direct article-thesis replication.",
            "The just-transition case tests the article's policy recommendation: recycle 30% and add efficiency measures.",
            "Gross tax proceeds are transfers to government, not social welfare gains; climate and health benefits are not monetized.",
            "The GDP dollar impact uses 2024 nominal GDP only as a scale proxy; it does not convert the IPSS model into a 2024 forecast.",
        ]),
        (19, "3. KEY LIMITATIONS", [
            "The underlying IPSS-UNDP research tables and sector elasticities were not publicly indexed at build time.",
            "Price response, recycling GDP offset, efficiency pass-through, leakage, and feasibility scores are explicit analyst assumptions.",
            "Sector incidence allocates national reduction using response weights; it is not a sector general-equilibrium model.",
            "No distributional microdata, commodity price pass-through, health co-benefits, or carbon-border effects are modeled.",
            "Avoid adding the mechanical fiscal revenue to GDP benefits; doing so would double-count a transfer.",
        ]),
        (27, "4. DECISION USE", [
            "Use for policy option screening, stakeholder discussion, and identifying the inputs that require primary research.",
            "Do not use as a statutory revenue forecast, budget score, or final welfare analysis without the underlying CGE model and administrative tax base.",
        ]),
    ]
    for start, heading, bullets in sections:
        section(ws, start, heading, 10)
        for i, bullet in enumerate(bullets, start + 1):
            ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=10)
            ws.cell(i, 1, "• " + bullet)
            ws.cell(i, 1).alignment = Alignment(wrap_text=True, vertical="top")
            ws.row_dimensions[i].height = 24
    section(ws, 33, "5. VERSION HISTORY", 10)
    header_row(ws, 34, ["Version", "Date", "Change", "Author"])
    ws.append(["v1.0", date(2026, 7, 14), "Initial five-case, three-scenario thesis simulation", "Codex"])
    ws["B35"].number_format = "dd mmm yyyy"
    widths = {"A": 22, "B": 16, "C": 54, "D": 18, "E": 14, "F": 14, "G": 14, "H": 14, "I": 14, "J": 14}
    configure_sheet(ws, widths, "A4")


def write_data_inventory() -> None:
    fields = ["metric_name", "value", "unit", "source_type", "source_title", "citation_url", "confidence", "date_range", "category", "notes"]
    rows = []
    for name, category, value, unit, source_type, source_title, url, confidence, notes, _key in ASSUMPTIONS:
        rows.append({
            "metric_name": name,
            "value": value,
            "unit": unit,
            "source_type": source_type,
            "source_title": source_title,
            "citation_url": url,
            "confidence": confidence,
            "date_range": "2030 BAU" if "emissions" in name.lower() or "ndc" in name.lower() else "2024-2026",
            "category": category,
            "notes": notes,
        })
    with DATA_INVENTORY.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fields)
        writer.writeheader()
        writer.writerows(rows)


def build() -> None:
    wb = Workbook()
    wb.remove(wb.active)
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = "auto"

    build_cover(wb)
    build_assumptions(wb)
    case_rows = build_scenario_comparison(wb)
    global ENGINE_ROWS
    ENGINE_ROWS = build_engine(wb, case_rows)
    build_executive_summary(wb)
    # Reorder executive sheets before detail sheets.
    wb._sheets = [
        wb["Cover"], wb["Executive Summary"], wb["Scenario Comparison"],
        wb.create_sheet("Decision Matrix Placeholder"), wb["Assumptions Log"], wb["Revenue Build-Up"]
    ] + [ws for ws in wb.worksheets if ws.title not in {"Cover", "Executive Summary", "Scenario Comparison", "Decision Matrix Placeholder", "Assumptions Log", "Revenue Build-Up"}]
    # Replace temporary slot with the real decision matrix while preserving order.
    placeholder = wb["Decision Matrix Placeholder"]
    wb.remove(placeholder)
    build_decision_matrix(wb)
    decision = wb["Decision Matrix"]
    wb._sheets.remove(decision)
    wb._sheets.insert(3, decision)

    build_fiscal_social(wb)
    build_macro(wb)
    build_roi_ndc(wb)
    build_sector_incidence(wb)
    build_sensitivity(wb)
    build_sources(wb)
    build_methodology(wb)

    target_order = [
        "Cover", "Executive Summary", "Scenario Comparison", "Decision Matrix", "Assumptions Log",
        "Revenue Build-Up", "Cost & Investment", "P&L Waterfall", "ROI & Payback",
        "Sensitivity Analysis", "Data Sources", "Methodology Notes", "Sector Incidence",
    ]
    wb._sheets = [wb[name] for name in target_order]
    for idx, ws in enumerate(wb.worksheets):
        ws.sheet_properties.tabColor = DARK if idx == 0 else (DARK_GREEN if idx <= 3 else BLUE)
        if idx <= 3:
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 1

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT)
    write_data_inventory()
    print(OUTPUT)


if __name__ == "__main__":
    build()
